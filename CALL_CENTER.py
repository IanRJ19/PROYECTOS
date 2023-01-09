from tkinter import *
from tkinter import filedialog
import os 
import openpyxl as ox

def seleccionar_carpeta():
    global ruta
    ruta = filedialog.askdirectory()
    ruta = str(ruta)
    print(ruta)

    if ruta:
        notif.config(fg="green", text="Carpeta Seleccionada")
        if not os.path.exists(ruta+"/"+"RESULTADOS"):
            os.makedirs(ruta+"/"+"RESULTADOS")
    else:
        notif.config(fg="red", text="Carpeta No Seleccionada")

def ACTUALIZAR_BD():
    try:
        import pandas as pd
        import time
        import os
        import time
        start_time = time.time()
        
        #Lectura de archivos
        archivos=os.listdir(ruta)

        for i in range(len(archivos)):
                    a=archivos[i]
                    if ("Detalle" in a):
                        if ("Macro" in a):
                            print(a)
                            b=pd.read_excel(ruta+"/"+a,skiprows=4,dtype=str,sheet_name="Detalle")
                            df_BBVA_M.append(b)

                        else:
                            print(a)
                            x=pd.read_excel(ruta+"/"+a,dtype=str)
                            probando=x.iloc[6,0]
        def update_spreadsheet(path:str ="C:/Users/Rayzek/Desktop/CALL CENTER/Reporte Indicadores Mensual_202212_.xlsx", _df=CF, startcol:int=1, startrow:int=1, sheet_name:str ="TDSheet"):
            wb = ox.load_workbook(path)
            ws=wb[sheet_name]
            for row in range(0, _df.shape[0]): #For each row in the dataframe
                for col in range(0, _df.shape[1]): #For each column in the dataframe
                    ws.cell(row = startrow + row, column = startcol + col).value = _df.iat[row, col]
            wb.save(path)

        update_spreadsheet("C:/Users/Rayzek/Desktop/CALL CENTER/Reporte Indicadores Mensual_202212_.xlsx", CF, 1, 2, "BD") 
        

        #Imprimimos tiempo
        tiempo=round((time.time() - start_time),2)
        texto="Proceso completado, tomó "+str(tiempo)+" segundos"
        #text_var = tkinter.StringVar(value=texto)
        #label_1 = customtkinter.CTkLabel(master=ventana, justify=tkinter.LEFT,textvariable=text_var,text_font=("Calibri",14,"bold"),text_color="#007E06")
        #label_1.pack(pady=12, padx=10)
        notif.config(fg="green", text=texto)

    except Exception:
            if "ruta" in globals():
                notif.config(fg="red", text="Error en el proceso")
            else:
                notif.config(fg="red", text="No ha seleccionado una carpeta")

def ACTUALIZAR_BDS():
    try:
        import pandas as pd
        import time
        import os
        import time
        start_time = time.time()
        
        #Lectura de archivos
        archivos=os.listdir(ruta)

        for i in range(len(archivos)):
                    a=archivos[i]
                    if ("Detalle" in a):
                        if ("Macro" in a):
                            print(a)
                            b=pd.read_excel(ruta+"/"+a,skiprows=4,dtype=str,sheet_name="Detalle")
                            df_BBVA_M.append(b)

                        else:
                            print(a)
                            x=pd.read_excel(ruta+"/"+a,dtype=str)
                            probando=x.iloc[6,0]

        def update_spreadsheet(path:str ="C:/Users/Rayzek/Desktop/CALL CENTER/Reporte Indicadores Mensual_202212_.xlsx", _df=CF, startcol:int=1, startrow:int=1, sheet_name:str ="TDSheet"):
            wb = ox.load_workbook(path)
            ws=wb[sheet_name]
            for row in range(0, _df.shape[0]): #For each row in the dataframe
                for col in range(0, _df.shape[1]): #For each column in the dataframe
                    ws.cell(row = startrow + row, column = startcol + col).value = _df.iat[row, col]
            wb.save(path)

        update_spreadsheet("C:/Users/Rayzek/Desktop/CALL CENTER/Reporte Indicadores Mensual_202212_.xlsx", da, 1, 2, "BDS") 

        #Imprimimos tiempo
        tiempo=round((time.time() - start_time),2)
        texto="Proceso completado, tomó "+str(tiempo)+" segundos"
        #text_var = tkinter.StringVar(value=texto)
        #label_1 = customtkinter.CTkLabel(master=ventana, justify=tkinter.LEFT,textvariable=text_var,text_font=("Calibri",14,"bold"),text_color="#007E06")
        #label_1.pack(pady=12, padx=10)
        notif.config(fg="green", text=texto)

    except Exception:
        if "ruta" in globals():
            notif.config(fg="red", text="Error en el proceso")
        else:
            notif.config(fg="red", text="No ha seleccionado una carpeta")
                        
###############

###############
# Pantalla Principal
master = Tk()
master.title("Inteligencia Comercial")
master.geometry("370x300")

# Etiquetas
Label(master, text="Automatización Call Center", fg="black", font=("Calibri", 18,"bold")).grid(padx=40,row=0,column=1,pady=20)


notif = Label(master, font=("Calibri", 14,"bold"))
notif.grid(sticky=N, pady=20, row=12, column=1)
notif.config(fg="green", text="Inicio")

# Botones

Button(master, width=20,fg="white", text="Actualizar BD",bg="#002060", font=("Calibri", 12,"bold"), command=ACTUALIZAR_BD).grid(row=9, column=1,pady=10,padx=0)

Button(master, width=20,fg="white", text="Actualizar BDS",bg="#002060", font=("Calibri", 12,"bold"), command=ACTUALIZAR_BDS).grid(row=10, column=1,pady=10,padx=0)

Button(master, width=20,fg="white", text="Seleccionar Carpeta",bg="#002060", font=("Calibri", 12,"bold"), command=seleccionar_carpeta).grid(row=11, column=1,pady=10,padx=60)

master.mainloop()


